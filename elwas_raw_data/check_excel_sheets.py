import openpyxl

wb = openpyxl.load_workbook('C:/Users/user/.gemini/scratch/elwas_export_D_ren.xlsx')
print(f"Sheet names: {wb.sheetnames}")
ws = wb.active
print(f"Dimensions: {ws.dimensions}")
for r in range(1, 10):
    row_vals = [ws.cell(r, c).value for c in range(1, 15)]
    print(f"Row {r}: {row_vals}")
